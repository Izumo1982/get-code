import pandas
import openpyxl
import glob
import os

get_dir = []
put_filename = "xxxx.xlsx"#ここに出力先のファイル名を入れます。（.xlsx　.csv　等）
get_dir.append("C:\\xxxx\\xxxx\\**\\*.xxx")#ここに、取得したいディレクトリ名を入れます。複数対応可（のはず）です。

for run_dir in get_dir:
    files = glob.glob(run_dir , recursive=True)
    li = []

    for file in files:
        print(str(file) + "ファイル名取得中")
        try:
            with open(file, mode='rt', encoding='utf-8') as f:
                read_data = f.readlines()
                for j in read_data:
                    li.append([file,j])
        except:
            pass

    book = openpyxl.Workbook()
    sheet = book.worksheets[0]
    sheet.cell(row= 1, column= 1).value = "ディレクトリ名　及び　ファイル名"
    sheet.cell(row= 1, column= 2).value = "内容"

for y, row in enumerate(li):
    for x, cell in enumerate(row):
        sheet.cell(row= y + 2, column= x + 1).value = li[y][x]
    if y % 10000 == 0: print(str(y) + "行終了")
book.save(put_filename)
print("対象のファイル　すべて終了")